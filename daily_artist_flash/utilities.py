# from daily_artist_flash.queries import select_queries
from datetime import datetime
import snowflake.connector
import pandas as pd
from dotenv import load_dotenv
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# environment variables from file
load_dotenv()
# readable print out tables
desired_width = 320
pd.set_option("display.width", desired_width)
pd.set_option("display.max_columns", 12)


def fetch_data_as_df(cur, sql):
    cur.execute(sql)
    rows = 0
    df = pd.DataFrame()
    columns = [i[0] for i in cur.description]

    while True:
        dat = cur.fetchmany(50000)
        if not dat:
            break
        temp = pd.DataFrame(dat, columns=columns)
        rows += df.shape[0]

        df = df.append(temp)

    # logging.info('%s rows extracted from Snowflake' % (rows))

    return df


def save_query_data_snowflake(cursor, query, f_name):
    data_df = fetch_data_as_df(cursor, query)
    output_fp = '%s_%s' % (f_name, datetime.today().strftime('%Y-%m-%d'))
    data_df.to_csv(output_fp, index=False)


def connect_to_snowflake():
    USER_SNOW = os.getenv("USER_SNOW")
    PASSWORD = os.getenv("PASSWORD")
    AUTHENTICATOR = os.getenv("AUTHENTICATOR")
    ACCOUNT = os.getenv("ACCOUNT")
    WAREHOUSE = os.getenv("WAREHOUSE")
    DATABASE = os.getenv("DATABASE")
    SCHEMA = os.getenv("SCHEMA")

    ctx = snowflake.connector.connect(user=USER_SNOW, password=PASSWORD,
                                      authenticator=AUTHENTICATOR, account=ACCOUNT,
                                      warehouse=WAREHOUSE, database=DATABASE, schema=SCHEMA)

    cursor = ctx.cursor()

    return cursor


def get_query_dfs(selected_queries, cursor):
    hot_hits_uk_df = fetch_data_as_df(cursor, selected_queries[0])
    todays_hits_apple_uk_df = fetch_data_as_df(cursor, selected_queries[1])
    todays_top_hits_spotify_df = fetch_data_as_df(cursor, selected_queries[2])
    spotify_daily_top_200_gb_df = fetch_data_as_df(cursor, selected_queries[3])
    query_total_streams_dsp_df = fetch_data_as_df(cursor, selected_queries[4])
    spotify_daily_top_200_ww_df = fetch_data_as_df(cursor, selected_queries[5])
    apple_music_daily_top_100_gb_df = fetch_data_as_df(cursor, selected_queries[6])
    shazam_top_200_gb_df = fetch_data_as_df(cursor, selected_queries[7])
    shazam_top_200_ww_df = fetch_data_as_df(cursor, selected_queries[8])
    occ_top_100_df = fetch_data_as_df(cursor, selected_queries[9])

    return hot_hits_uk_df, todays_hits_apple_uk_df, todays_top_hits_spotify_df, spotify_daily_top_200_gb_df, query_total_streams_dsp_df, \
           spotify_daily_top_200_ww_df, apple_music_daily_top_100_gb_df, shazam_top_200_gb_df, shazam_top_200_ww_df, occ_top_100_df


def week_change_calculator(df_week_prior, df_week):
    if (df_week != 'N/A') & (df_week_prior != 'N/A'):
        df_week_prior = df_week - df_week_prior
    elif (df_week == 'N/A') & (df_week_prior != 'N/A'):
        df_week_prior = -df_week_prior
    elif (df_week != 'N/A') & (df_week_prior == 'N/A'):
        df_week_prior = df_week
    elif df_week == df_week_prior:
        df_week_prior = 0
    elif (df_week == 'N/A') & (df_week_prior == 'N/A'):
        df_week_prior = 0
    else:
        df_week_prior = 'err'
    return df_week_prior


def hot_hits_uk(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
        df['PLAYLIST_NAME'] = df['PLAYLIST_NAME'].str.replace('\'', '', regex=True)
        df = df.loc[(df['PLAYLIST_NAME'] == playlist)]
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['PLAYLIST_NAME'] == playlist)]
        df_week_prior = df_week_prior['TRACK_POSITION'].values[0]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['PLAYLIST_NAME'] == playlist)]
        df_week = df_week['TRACK_POSITION'].values[0]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def todays_top_hits(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
        df['PLAYLIST_NAME'] = df['PLAYLIST_NAME'].str.replace('\'', '', regex=True)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['PLAYLIST_NAME'] == playlist)]
        df_week_prior = df_week_prior['TRACK_POSITION'].values[0]

    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['PLAYLIST_NAME'] == playlist)]
        df_week = df_week['TRACK_POSITION'].values[0]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def todays_hits(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
        df['PLAYLIST_NAME'] = df['PLAYLIST_NAME'].str.replace('\'', '', regex=True)
        df = df.loc[(df['PLAYLIST_NAME'] == playlist)]
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['PLAYLIST_NAME'] == playlist)]
        df_week_prior = df_week_prior['TRACK_POSITION'].values[0]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['PLAYLIST_NAME'] == playlist)]
        df_week = df_week['TRACK_POSITION'].values[0]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def spotify_daily_200_gb_selector(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 2]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 2]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def youtube_streams_summed(df, week_prior, week):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == 'YouTube')]
        df_week_prior = df_week_prior['SUMMED_STREAMS'].sum()
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == 'YouTube')]
        df_week = df_week['SUMMED_STREAMS'].sum()
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def spotify_top_200_global(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def apple_music_daily_top_100_gb(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def shazam_top_200_gb(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def shazam_top_200_ww(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def occ_top_100(df, week_prior, week, playlist):
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week, df_week_prior)

    return df_week_prior, df_week


def create_row(dfs, week_prior, week):
    data_row = []
    week_selector = [0, 1]
    data_point_gatherer = [
        hot_hits_uk(dfs[0], week_prior, week, "Hot Hits UK"),
        todays_top_hits(dfs[1], week_prior, week, "Todays Hits"),
        todays_hits(dfs[2], week_prior, week, "Todays Top Hits"),
        spotify_daily_200_gb_selector(dfs[3], week_prior, week, "Top 200"),
        youtube_streams_summed(dfs[4], week_prior, week),
        spotify_top_200_global(dfs[5], week_prior, week, "Top 200"),
        apple_music_daily_top_100_gb(dfs[6], week_prior, week, "Top Songs"),
        shazam_top_200_gb(dfs[7], week_prior, week, "SHAZAM TOP 200"),
        shazam_top_200_ww(dfs[8], week_prior, week, "SHAZAM TOP 200"),
        occ_top_100(dfs[9], week_prior, week, "Top 100 Combined Singles")
    ]

    for x in data_point_gatherer:
        for y in week_selector:
            data_row.append(x[y])

    return data_row


def data_row_dict(row_data):
    data_row = {"Hot Hits UK (Spotify)": {"week prior change": row_data[0], "current week": row_data[1]},
                "Today\'s Top Hits (Spotify)": {"week prior change": row_data[2], "current week": row_data[3]},
                "Today\'s Hits (Apple)": {"week prior change": row_data[4], "current week": row_data[5]},
                "Spotify Daily Top 200 (GB)": {"week prior change": row_data[6], "current week": row_data[7]},
                "Youtube Views (Global)": {"week prior change": row_data[8], "current week": row_data[9]},
                "Spotify Daily Top 200 (Global)": {"week prior change": row_data[10], "current week": row_data[11]},
                "Apple Music Daily Top 100 (GB)": {"week prior change": row_data[12], "current week": row_data[13]},
                "Shazam Top 200 (GB)": {"week prior change": row_data[14], "current week": row_data[15]},
                "Shazam Top 200 (Global)": {"week prior change": row_data[16], "current week": row_data[17]},
                "OCC Top 100 Singles": {"week prior change": row_data[18], "current week": row_data[19]}
                }
    return data_row


def create_daily_flash():
    index = pd.Index([], name='artist - song')

    columns = pd.MultiIndex.from_product(
        [['Hot Hits UK (Spotify)', 'Today\'s Top Hits (Spotify)', 'Today\'s Hits (Apple)',
          'Spotify Daily Top 200 (GB)', 'Youtube Views (Global)', 'Spotify Daily Top 200 (Global)',
          'Apple Music Daily Top 100 (GB)', 'Shazam Top 200 (GB)', 'Shazam Top 200 (Global)',
          'OCC Top 100 Singles'],
         ['week prior change', 'current week']], names=['source', 'week'])

    daily_flash = pd.DataFrame(columns=columns, index=index)

    return daily_flash


def add_row_to_daily_flash(data_row, daily_flash, artist, track_title):
    artist_track = artist + ' - ' + track_title
    daily_flash = daily_flash.append(pd.DataFrame.from_dict(data_row).unstack().rename(f"{artist_track}"))

    return daily_flash


def daily_flash_to_excel(daily_flash_df, artist_track_dict, week):
    row_len = len(artist_track_dict)
    # save work into excel format
    wb = Workbook()
    ws = wb.create_sheet('Daily Flash')
    wb.active = ws

    rows = dataframe_to_rows(daily_flash_df, index=True, header=True)
    for r in rows:
        ws.append(r)

    font = Font(bold=True)
    red_fill = PatternFill(patternType='solid', fgColor='f4cccc')
    green_fill = PatternFill(patternType='solid', fgColor='d9ead3')

    # header rows bold
    for col in range(1, 24):
        for row in range(1, 3):
            ws.cell(row=row, column=col).font = font

    # track rows bold
    for col in range(1, 2):
        for row in range(1, 4 + row_len):
            ws.cell(row=row, column=col).font = font

    sheet = wb['Daily Flash']

    # red or green fill
    # add in len dictionary of artists
    for col in range(2, 24, 2):
        for row in range(1, 4 + row_len):
            if type(sheet.cell(row=row, column=col + 1).value) == int:
                if sheet.cell(row=row, column=col).value < 0:
                    sheet.cell(row=row, column=col + 1).fill = green_fill
                elif sheet.cell(row=row, column=col).value > 0:
                    sheet.cell(row=row, column=col + 1).fill = red_fill
                elif sheet.cell(row=row, column=col).value == 0:
                    pass
                elif sheet.cell(row=row, column=col).value == 'N/A':
                    pass

    # Youtube Exception
    for col in range(10, 11):
        for row in range(1, 4 + row_len):
            if type(sheet.cell(row=row, column=col + 1).value) == int:
                if sheet.cell(row=row, column=col).value < 0:
                    sheet.cell(row=row, column=col + 1).fill = red_fill
                elif sheet.cell(row=row, column=col).value > 0:
                    sheet.cell(row=row, column=col + 1).fill = green_fill
                elif sheet.cell(row=row, column=col).value == 0:
                    pass
                elif sheet.cell(row=row, column=col).value == 'N/A':
                    pass

    title = f"daily_flash_{week}.xlsx"
    wb.save(title)

    return
