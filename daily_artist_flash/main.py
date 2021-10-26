from daily_artist_flash.utilities import connect_to_snowflake, get_query_dfs, week_change_calculator, \
    hot_hits_uk, todays_top_hits, todays_hits, spotify_daily_200_gb_selector, youtube_streams_summed, \
    create_row, data_row_dict, create_daily_flash, add_row_to_daily_flash, fetch_data_as_df

from daily_artist_flash.queries import select_queries

from dotenv import load_dotenv
import os
import datetime
import snowflake.connector
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill


def main():
    # load environment variables
    load_dotenv()
    # connect to snowflake
    cursor = connect_to_snowflake()
    # create daily flash df
    daily_flash = create_daily_flash()

    # artists & tracks to check
    artist_track_dict = {
        'Clean Bandit x Topic': 'Drive (feat. Wes Nelson)',
        'Ed Sheeran': 'Bad Habits'
    }
    week_prior = '2021-09-17'
    week = '2021-09-24'

    # iterate through utility functions and collect data
    for k, v in artist_track_dict.items():
        selected_queries = select_queries(2021, 9, 24, k, v)
        query_dfs = get_query_dfs(selected_queries, cursor)
        data_rows = create_row(query_dfs, week_prior, week)
        data_row = data_row_dict(data_rows)
        daily_flash = add_row_to_daily_flash(data_row, daily_flash, k, v)

    print(daily_flash)

    # save work into excel format
    wb = Workbook()
    ws = wb.create_sheet('Daily Flash')

    rows = dataframe_to_rows(daily_flash, index=True, header=True)
    for r in rows:
        ws.append(r)

    # for row in ws["1:1"]:
    #     for cell in row:
    #         cell.font = Font(b=True)

    # for cell in ws["1:2"]:


    title = f"daily_flash_{week}.xlsx"
    wb.save(title)






if __name__ == "__main__":
    main()

