
from datetime import datetime, date
import snowflake.connector
import pandas as pd
from dotenv import load_dotenv
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Side
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from utils.queries import fact_audio_playlist_track_metrics_latest_ingest, fact_charts_daily_latest_ingest, \
                        fact_charts_weekly_latest_ingest, fact_audio_streaming_latest_ingest, fact_audio_streaming_agg_weekly_ingest

# environment variables from file
load_dotenv()
# readable print out tables for testing purposes
desired_width = 320
pd.set_option("display.width", desired_width)
pd.set_option("display.max_columns", 12)


def fetch_data_as_df(cur, sql):
    cur.execute(sql)
    rows = 0
    df = pd.DataFrame()
    columns = [i[0] for i in cur.description]

    while True:
        dat = cur.fetchmany(50000)
        if not dat:
            break
        temp = pd.DataFrame(dat, columns=columns)
        rows += df.shape[0]

        df = df.append(temp)

    # logging.info('%s rows extracted from Snowflake' % (rows))

    return df


def save_query_data_snowflake(cursor, query, f_name):
    data_df = fetch_data_as_df(cursor, query)
    output_fp = '%s_%s' % (f_name, datetime.today().strftime('%Y-%m-%d'))
    data_df.to_csv(output_fp, index=False)


def connect_to_snowflake():
    USER_SNOW = os.getenv("USER_SNOW")
    PASSWORD = os.getenv("PASSWORD")
    AUTHENTICATOR = os.getenv("AUTHENTICATOR")
    ACCOUNT = os.getenv("ACCOUNT")
    WAREHOUSE = os.getenv("WAREHOUSE")
    DATABASE = os.getenv("DATABASE")
    SCHEMA = os.getenv("SCHEMA")

    ctx = snowflake.connector.connect(user=USER_SNOW, password=PASSWORD,
                                      authenticator=AUTHENTICATOR, account=ACCOUNT,
                                      warehouse=WAREHOUSE, database=DATABASE, schema=SCHEMA)

    cursor = ctx.cursor()

    return cursor


def get_week_week_prior_dates(up_to_max_date_query):
    cursor = connect_to_snowflake()
    date = fetch_data_as_df(cursor, up_to_max_date_query)
    week = date['DATE'][0]
    delta = timedelta(weeks=1)
    week_prior = str(week - delta)
    week = str(week)
    
    return week_prior, week


def get_week_prior_dates_and_week_dates_from_sflake_ingests():
    fact_audio_playlist_track_metrics_week_prior, fact_audio_playlist_track_metrics_week = get_week_week_prior_dates(fact_audio_playlist_track_metrics_latest_ingest)
    fact_charts_daily_week_prior, fact_charts_daily_week = get_week_week_prior_dates(fact_charts_daily_latest_ingest)
    fact_charts_weekly_week_prior, fact_charts_weekly_week = get_week_week_prior_dates(fact_charts_weekly_latest_ingest)
    fact_audio_streaming_week_prior, fact_audio_streaming_week = get_week_week_prior_dates(fact_audio_streaming_latest_ingest)
    fact_audio_streaming_agg_weekly_week_prior, fact_audio_streaming_agg_weekly_week = get_week_week_prior_dates(fact_audio_streaming_agg_weekly_ingest)


    return fact_audio_playlist_track_metrics_week_prior, fact_audio_playlist_track_metrics_week, \
            fact_charts_daily_week_prior, fact_charts_daily_week, \
            fact_charts_weekly_week_prior, fact_charts_weekly_week, \
            fact_audio_streaming_week_prior, fact_audio_streaming_week, \
            fact_audio_streaming_agg_weekly_week_prior, fact_audio_streaming_agg_weekly_week

week_prior_weeks = get_week_prior_dates_and_week_dates_from_sflake_ingests()


def get_query_dfs(selected_queries, cursor):
    hot_hits_uk_df = fetch_data_as_df(cursor, selected_queries[0])
    todays_hits_apple_uk_df = fetch_data_as_df(cursor, selected_queries[1])
    todays_top_hits_spotify_df = fetch_data_as_df(cursor, selected_queries[2])
    spotify_daily_top_200_gb_df = fetch_data_as_df(cursor, selected_queries[3])
    query_total_streams_dsp_df = fetch_data_as_df(cursor, selected_queries[4])
    spotify_daily_top_200_ww_df = fetch_data_as_df(cursor, selected_queries[5])
    apple_music_daily_top_100_gb_df = fetch_data_as_df(cursor, selected_queries[6])
    shazam_top_200_gb_df = fetch_data_as_df(cursor, selected_queries[7])
    shazam_top_200_ww_df = fetch_data_as_df(cursor, selected_queries[8])
    occ_top_100_singles_df = fetch_data_as_df(cursor, selected_queries[9])
    dsp_streams_df = fetch_data_as_df(cursor, selected_queries[10])
    youtube_ugc_pgc_df = fetch_data_as_df(cursor, selected_queries[11])

    
    
    return hot_hits_uk_df, todays_hits_apple_uk_df, todays_top_hits_spotify_df, spotify_daily_top_200_gb_df, query_total_streams_dsp_df, \
           spotify_daily_top_200_ww_df, apple_music_daily_top_100_gb_df, shazam_top_200_gb_df, shazam_top_200_ww_df, occ_top_100_singles_df, dsp_streams_df, \
               youtube_ugc_pgc_df


def week_change_calculator(df_week_prior, df_week):
    if (df_week != 'N/A') & (df_week_prior != 'N/A'):
        df_week_prior = df_week - df_week_prior
    elif (df_week == 'N/A') & (df_week_prior != 'N/A'):
        df_week_prior = -df_week_prior
    elif (df_week != 'N/A') & (df_week_prior == 'N/A'):
        df_week_prior = df_week
    elif df_week == df_week_prior:
        df_week_prior = 0
    elif (df_week == 'N/A') & (df_week_prior == 'N/A'):
        df_week_prior = 0
    else:
        df_week_prior = 'err'
    return df_week_prior


def hot_hits_uk(df, playlist):

    week_prior = week_prior_weeks[0]
    week = week_prior_weeks[1]

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
        df = df.loc[(df['PLAYLIST_NAME'] == playlist)]
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['PLAYLIST_NAME'] == playlist)]
        df_week_prior = df_week_prior['TRACK_POSITION'].values[0]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['PLAYLIST_NAME'] == playlist)]
        df_week = df_week['TRACK_POSITION'].values[0]
        
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def todays_top_hits(df, playlist):

    week_prior = week_prior_weeks[0]
    week = week_prior_weeks[1]
    
    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
        df['PLAYLIST_NAME'] = df['PLAYLIST_NAME'].str.replace('\'', '', regex=True)
        df = df.loc[(df['PLAYLIST_NAME'] == playlist)]
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['PLAYLIST_NAME'] == playlist)]
        df_week_prior = df_week_prior['TRACK_POSITION'].values[0]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['PLAYLIST_NAME'] == playlist)]
        df_week = df_week['TRACK_POSITION'].values[0]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def todays_hits(df, playlist):

    week_prior = date.fromisoformat(week_prior_weeks[0])
    week = date.fromisoformat(week_prior_weeks[1])
    delta = timedelta(days=1)
    week_prior = str(week_prior - delta)
    week = str(week - delta) 

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
        df['PLAYLIST_NAME'] = df['PLAYLIST_NAME'].str.replace('\'', '', regex=True)
        df = df.loc[(df['PLAYLIST_NAME'] == playlist)]
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['PLAYLIST_NAME'] == playlist)]
        df_week_prior = df_week_prior['TRACK_POSITION'].values[0]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['PLAYLIST_NAME'] == playlist)]
        df_week = df_week['TRACK_POSITION'].values[0]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def spotify_daily_200_gb_selector(df, playlist):

    week_prior = week_prior_weeks[2]
    week = week_prior_weeks[3]   

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 2]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 2]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def youtube_streams_summed(df):

    week_prior = week_prior_weeks[6]
    week = week_prior_weeks[7]

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == 'YouTube')]
        df_week_prior = df_week_prior['SUMMED_STREAMS'].sum()
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == 'YouTube')]
        df_week = df_week['SUMMED_STREAMS'].sum()
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def spotify_top_200_global(df, playlist):

    week_prior = week_prior_weeks[2]
    week = week_prior_weeks[3]  

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def apple_music_daily_top_100_gb(df, playlist):

    # Delay in the data reporting
    week_prior = date.fromisoformat(week_prior_weeks[2])
    week = date.fromisoformat(week_prior_weeks[3])
    delta = timedelta(days=3)
    week_prior = str(week_prior - delta)
    week = str(week - delta)

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def shazam_top_200_gb(df, playlist):

    week_prior = week_prior_weeks[4]
    week = week_prior_weeks[5] 

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def shazam_top_200_ww(df, playlist):

    week_prior = week_prior_weeks[4]
    week = week_prior_weeks[5] 

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def occ_top_100_singles(df, playlist):

    # Weekly chart ingest on Friday, run on Monday with back date. delta = 3 if run on a Mon
    week_prior = date.fromisoformat(week_prior_weeks[4])
    week = date.fromisoformat(week_prior_weeks[5])
    delta = timedelta(days=0)
    week_prior = str(week_prior - delta)
    week = str(week - delta)

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CHART_NAME'] == playlist)]
        df_week_prior = df_week_prior.iloc[0, 3]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CHART_NAME'] == playlist)]
        df_week = df_week.iloc[0, 3]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week

def dsp_streams_uk(df, customer_name):

    # Weekly chart ingest on Friday, run on Monday with back date. delta = 2 if run on a Mon
    week_prior = date.fromisoformat(week_prior_weeks[8])
    week = date.fromisoformat(week_prior_weeks[9])
    delta = timedelta(days=0)
    week_prior = str(week_prior - delta)
    week = str(week - delta)

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week_prior = df_week_prior.iloc[0, 4]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week = df_week.iloc[0, 4]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week

def dsp_streams_global(df, customer_name):

    # Weekly chart ingest on Friday, run on Monday with back date. delta = 2 if run on a Mon
    week_prior = date.fromisoformat(week_prior_weeks[8])
    week = date.fromisoformat(week_prior_weeks[9])
    delta = timedelta(days=0)
    week_prior = str(week_prior - delta)
    week = str(week - delta)

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week_prior = df_week_prior.iloc[0, 5]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week = df_week.iloc[0, 5]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week

def dsp_streams_total(df, customer_name):

    # Weekly chart ingest on Friday, run on Monday with back date. delta = 2 if run on a Mon
    week_prior = date.fromisoformat(week_prior_weeks[8])
    week = date.fromisoformat(week_prior_weeks[9])
    delta = timedelta(days=0)
    week_prior = str(week_prior - delta)
    week = str(week - delta)

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week_prior = df_week_prior.iloc[0, 5]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week = df_week.iloc[0, 5]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week

def dsp_streams_total(df, customer_name):

    # Weekly chart ingest on Friday, run on Monday with back date. delta = 2 if run on a Mon
    week_prior = '2022-01-20'
    week = '2022-01-27'

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week_prior = df_week_prior.iloc[0, 4] + df_week_prior.iloc[0, 5]

    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week = df_week.iloc[0, 4] + df_week.iloc[0, 5]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week

def youtube_ugc_pgc_streams(df, customer_name):

    # Weekly chart ingest on Friday, run on Monday with back date. delta = 2 if run on a Mon
    week_prior = date.fromisoformat(week_prior_weeks[8])
    week = date.fromisoformat(week_prior_weeks[9])
    delta = timedelta(days=0)
    week_prior = str(week_prior - delta)
    week = str(week - delta)

    try:
        df['DATE_KEY'] = df['DATE_KEY'].astype(str)
    except Exception:
        pass
    try:
        df_week_prior = df.loc[(df['DATE_KEY'] == week_prior) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week_prior = df_week_prior.iloc[0, 4]
    except Exception:
        df_week_prior = 'N/A'
    try:
        df_week = df.loc[(df['DATE_KEY'] == week) & (df['CUSTOMER_NAME'] == customer_name)]
        df_week = df_week.iloc[0, 4]
    except Exception:
        df_week = 'N/A'

    df_week_prior = week_change_calculator(df_week_prior, df_week)

    return df_week_prior, df_week


def create_row(dfs):
    data_row = []
    week_selector = [0, 1]
    data_point_gatherer = [
        hot_hits_uk(dfs[0], "Hot Hits UK"),
        todays_hits(dfs[1], "Éxitos de hoy"),
        todays_top_hits(dfs[2], "Todays Top Hits"),
        spotify_daily_200_gb_selector(dfs[3], "Top 200"),
        youtube_streams_summed(dfs[4]),
        spotify_top_200_global(dfs[5], "Top 200"),
        apple_music_daily_top_100_gb(dfs[6], "Top Songs"),
        shazam_top_200_gb(dfs[7], "SHAZAM TOP 200"),
        shazam_top_200_ww(dfs[8], "SHAZAM TOP 200"),
        occ_top_100_singles(dfs[9], "Top 100 Combined Singles"),
        dsp_streams_uk(dfs[10], "Spotify"),
        dsp_streams_global(dfs[10], "Spotify"),
        dsp_streams_total(dfs[10], "Spotify"),
        dsp_streams_uk(dfs[10], "Apple Music"),
        dsp_streams_global(dfs[10], "Apple Music"),
        dsp_streams_total(dfs[10], "Apple Music"),
        dsp_streams_uk(dfs[10], "Amazon Prime"),
        dsp_streams_global(dfs[10], "Amazon Prime"),
        dsp_streams_total(dfs[10], "Amazon Prime"),
        dsp_streams_uk(dfs[10], "Amazon Ad Supported"),
        dsp_streams_global(dfs[10], "Amazon Ad Supported"),
        dsp_streams_total(dfs[10], "Amazon Ad Supported"),
        dsp_streams_uk(dfs[10], "Amazon Unlimited"),
        dsp_streams_global(dfs[10], "Amazon Unlimited"),
        dsp_streams_total(dfs[10], "Amazon Unlimited"),
        youtube_ugc_pgc_streams(dfs[11], "YouTube Official"),
        youtube_ugc_pgc_streams(dfs[11], "YouTube UGC"),
    ]

    for x in data_point_gatherer:
        for y in week_selector:
            data_row.append(x[y])

    return data_row


def data_row_dict(row_data):
    data_row = {"Hot Hits UK (Spotify)": {"": row_data[0], "current week": row_data[1]},
               "Today\'s Hits (Apple)": {"": row_data[2], "current week": row_data[3]},
               "Today\'s Top Hits (Spotify)": {"": row_data[4], "current week": row_data[5]},
               "Spotify Daily Top 200 (GB)": {"": row_data[6], "current week": row_data[7]},
               "Youtube Views (Global)": {"": row_data[8], "current week": row_data[9]},
               "Spotify Daily Top 200 (Global)": {"": row_data[10], "current week": row_data[11]},
               "Apple Music Daily Top 100 (GB)": {"": row_data[12], "current week": row_data[13]},
               "Shazam Top 200 (GB)": {"": row_data[14], "current week": row_data[15]},
               "Shazam Top 200 (Global)": {"": row_data[16], "current week": row_data[17]},
               "OCC Top 100 Singles": {"": row_data[18], "current week": row_data[19]},
               "Spotify Weekly Streams (GB)": {"": row_data[20], "current week": row_data[21]},
               "Spotify Weekly Streams (Global)": {"": row_data[22], "current week": row_data[23]},
               "Spotify Weekly Streams (Total)": {"": row_data[24], "current week": row_data[25]},
               "Apple Music Weekly Streams (GB)": {"": row_data[26], "current week": row_data[27]},
               "Apple Music Weekly Streams (Global)": {"": row_data[28], "current week": row_data[29]},
               "Apple Music Weekly Streams (Total)": {"": row_data[30], "current week": row_data[31]},
               "Amazon Prime Weekly Streams (GB)": {"": row_data[32], "current week": row_data[33]},
               "Amazon Prime Weekly Streams (Global)": {"": row_data[34], "current week": row_data[35]},
               "Amazon Prime Weekly Streams (Total)": {"": row_data[36], "current week": row_data[37]},
               "Amazon Ad Supported Weekly Streams (GB)": {"": row_data[38], "current week": row_data[39]},
               "Amazon Ad Supported Weekly Streams (Global)": {"": row_data[40], "current week": row_data[41]},
               "Amazon Ad Supported Weekly Streams (Total)": {"": row_data[42], "current week": row_data[43]},
               "Amazon Unlimited Weekly Streams (GB)": {"": row_data[44], "current week": row_data[45]},
               "Amazon Unlimited Weekly Streams (Global)": {"": row_data[46], "current week": row_data[47]},
               "Amazon Unlimited Weekly Streams (Total)": {"": row_data[48], "current week": row_data[49]},
               "YouTube Official Streams (Total)": {"": row_data[50], "current week": row_data[51]},
               "YouTube UGC Streams (Total)": {"": row_data[52], "current week": row_data[53]},
               }
    return data_row



def create_daily_flash():
    index = pd.Index([], name='artist - song')

    columns = pd.MultiIndex.from_product(
        [['Hot Hits UK (Spotify)','Today\'s Hits (Apple)', 'Today\'s Top Hits (Spotify)',
          'Spotify Daily Top 200 (GB)', 'Youtube Views (Global)', 'Spotify Daily Top 200 (Global)',
          'Apple Music Daily Top 100 (GB)', 'Shazam Top 200 (GB)', 'Shazam Top 200 (Global)',
          'OCC Top 100 Singles', 'Spotify Weekly Streams (GB)', 'Spotify Weekly Streams (Global)', 'Spotify Weekly Streams (Total)',
          'Apple Music Weekly Streams (GB)', 'Apple Music Weekly Streams (Global)', 'Apple Music Weekly Streams (Total)',
          'Amazon Prime Weekly Streams (GB)', 'Amazon Prime Weekly Streams (Global)', 'Amazon Prime Weekly Streams (Total)',
          'Amazon Ad Supported Weekly Streams (GB)', 'Amazon Ad Supported Weekly Streams (Global)', 'Amazon Ad Supported Weekly Streams (Total)',
          'Amazon Unlimited Weekly Streams (GB)', 'Amazon Unlimited Weekly Streams (Global)', 'Amazon Unlimited Weekly Streams (Total)',
          'YouTube Official Streams (Total)', 'YouTube UGC Streams (Total)'],
          ['current week', '']], names=['source', 'week'])

    daily_flash = pd.DataFrame(columns=columns, index=index)

    return daily_flash


def add_row_to_daily_flash(data_row, daily_flash, artist, track_title):
    artist_track = artist + ' - ' + track_title
    daily_flash = daily_flash.append(pd.DataFrame.from_dict(data_row).unstack().rename(f"{artist_track}"))

    return daily_flash


def daily_flash_to_excel(daily_flash_df, artist_track_dict):

    week = datetime.now().date()
    row_len = len(artist_track_dict)
    # save work into excel format
    wb = Workbook()
    ws = wb.create_sheet('Daily Flash', 0)
    ws = wb.active
    
    # column A artist and track width wider
    for col in range(1,2):
        ws.column_dimensions[get_column_letter(col)].width = 50

    # rest of columns spacious tor title
    for col in range(2,57,2):
        ws.column_dimensions[get_column_letter(col)].width = 30

    for col in range(3,57,2):
        ws.column_dimensions[get_column_letter(col)].width = 10

    rows = dataframe_to_rows(daily_flash_df, index=True, header=True)
    for r in rows:
        ws.append(r)

    # cell formatting params
    font = Font(bold=True)
    red_fill = PatternFill(patternType='solid', fgColor='f4cccc')
    green_fill = PatternFill(patternType='solid', fgColor='d9ead3')

    side = Side(border_style='thin', color='00FFFFFF')
    no_border = Border(
    left=side, 
    right=side, 
    top=side, 
    bottom=side,
    )
    centered = Alignment(horizontal='center', vertical='center')

    # no border for cells and center text
    for col in range(1, 56):
        for row in range(1,row_len+4):
            ws.cell(row=row,column=col).border = no_border
            ws.cell(row=row,column=col).alignment = centered

    # header rows bold
    for col in range(1, 56):
        for row in range(1, 3):
            ws.cell(row=row, column=col).font = font

    # track rows bold
    for col in range(1, 2):
        for row in range(1, 4 + row_len):
            ws.cell(row=row, column=col).font = font

    sheet = wb['Daily Flash']

    # fill cell colours based on changes, -changes moving up chart is green, \
    # +changes going down chart red, no move no colour
    for col in range(2, 24, 2):
        for row in range(4, 4+row_len):
            if type(sheet.cell(row=row, column=col+1).value) == int:
                if sheet.cell(row=row, column=col+1).value < 0:
                    sheet.cell(row=row, column=col).fill = green_fill
                elif sheet.cell(row=row, column=col+1).value == 0:
                    pass
                elif sheet.cell(row=row, column=col+1).value > 0:
                    sheet.cell(row=row, column=col).fill = red_fill


    # Youtube & other DSPs streams colours reversed +changes good.
    for col in range(10, 11):
        for row in range(1, 4+row_len):
            if type(sheet.cell(row=row, column=col).value) == int:
                if sheet.cell(row=row, column=col+1).value < 0:
                    sheet.cell(row=row, column=col).fill = red_fill
                elif sheet.cell(row=row, column=col+1).value > 0:
                    sheet.cell(row=row, column=col).fill = green_fill
                elif sheet.cell(row=row, column=col).value == 0:
                    pass
                elif sheet.cell(row=row, column=col).value == 'N/A':
                    pass

    # Youtube & other DSPs streams colours reversed +changes good.
    for col in range(22, 56, 2):
        for row in range(1, 4+row_len):
            if type(sheet.cell(row=row, column=col).value) == int:
                if sheet.cell(row=row, column=col+1).value < 0:
                    sheet.cell(row=row, column=col).fill = red_fill
                elif sheet.cell(row=row, column=col+1).value > 0:
                    sheet.cell(row=row, column=col).fill = green_fill
                elif sheet.cell(row=row, column=col).value == 0:
                    pass
                elif sheet.cell(row=row, column=col).value == 'N/A':
                    pass
    
    #Add percentage change to streaming figs
    for col in range(10,11):
        for row in range(4, 17):
            z = sheet.cell(row=row, column=col+1)
            a = sheet.cell(row=row, column=col).coordinate
            b = sheet.cell(row=row, column=col+1).value
            n = f'=ROUND(IF({b}<0,({a}-({a}-{b}))/({a}-({b}))*100,(({b}/({a}-{b}))*100)),0)&CHAR(37)'
            z.value = n

    #Add percentage change to streaming figs
    for col in range(22, 56, 2):
        for row in range(4, 17):
            z = sheet.cell(row=row, column=col+1)
            a = sheet.cell(row=row, column=col).coordinate
            b = sheet.cell(row=row, column=col+1).value
            n = f'=ROUND(IF({b}<0,({a}-({a}-{b}))/({a}-({b}))*100,(({b}/({a}-{b}))*100)),0)&CHAR(37)'
            z.value = n

    #delete j column (youtube original), fix until new youtube query is deleted and code rearranged.
    sheet.delete_cols(idx=10)

    
    title = f"daily_flash_{week}.xlsx"
    wb.save(title)

    return
